# -*- coding: utf-8
#

import  os
import sys
import  shutil
def uu():
    from openpyxl import load_workbook
    from  openpyxl.reader.excel import load_workbook
    wb = load_workbook(filename = 'dic_schema.xlsx')
    sheet_ranges = wb['Sheet1']
    
    fo = open('html_dic.py', 'w')
    fo_edit = open('array_add_edit_view.py', 'w')
    list_fo = open('array_list.py', 'w')
    list_fo.write('''tlist_0100 = ['resolution', 'spatial', 'size']
tlist_0200 = ['subject', 'size']
tlist_0300 = ['size']
tlist_0400 = ['subject', 'size']
tlist_0500 = ['subject', 'size']
tlist_0600 = ['subject']
''')
    sig_name_arr = []
    for jj in ['D', 'E', 'F', 'G', 'H', 'I','J','K','L','M','N','O']:

        cc_val = sheet_ranges['{0}1'.format(jj)].value
        if  cc_val and cc_val != '':
            (qian,hou) = cc_val.split(':')
            c_name, e_name = qian.split(',')
            sig_name_arr.append(e_name)
            tags1 = hou.split(',')
            tags1 = [x.strip() for x in tags1]
            tags_dic = {}

            if len(tags1) == 1:
                tags_dic[1] = hou
                ctr_type = 'text'
            else:
                for ii in range(len(tags1)):
                    tags_dic[ii + 1] = tags1[ii].strip()
                ctr_type = 'select'
            print(tags1)

            fo.write('''html_{0} = {{
                'en': 'extra_{0}',
                'zh': '{1}',
                'dic': {2},
                'type': '{3}',
                }}\n'''.format( e_name, c_name, tags_dic,ctr_type))

    
    papa_index = 1
    c_index = 1
    papa_id  = 0
    uid = ''
    
    p_dic = {}


    for row_num in range(2, 50):
        if sheet_ranges['B{0}'.format(row_num)].value  and sheet_ranges['B{0}'.format(row_num)].value != '':
            c_index = 1
            # if papa_id:
            #     fo.write( 'p_dic_{0} = {1}\n'.format(papa_id, str(p_dic)))

            papa_id =  papa_index
            papa_index += 1
            print(papa_id)
            
            p_dic = {
                'uid': '0{0}'.format(papa_id),
                'name': sheet_ranges['B{0}'.format(row_num)].value,
                'u_arr': [],            
                }

        c_cell_val =  sheet_ranges['C{0}'.format(row_num)].value
        # print(c_cell_val)
        if  c_cell_val and c_cell_val != '' :
            c_id =  c_index
            print(c_index)
            u_dic = {}

            app_uid =  '0{0}0{1}'.format(papa_id, c_index)
            p_dic['u_arr'].append(app_uid)
            u_dic['uid'] = app_uid
            u_dic['name'] = sheet_ranges['C{0}'.format(row_num)].value
            u_dic['arr'] = []

            for ii,jj  in zip(['D', 'E', 'F', 'G', 'H', 'I','J','K','L','M','N','O'], sig_name_arr):
                cell_val = sheet_ranges['{0}{1}'.format(ii, row_num)].value
                if cell_val == 1:
                    u_dic['arr'].append('{0}'.format(jj))
            print(u_dic)
            fo_edit.write( 'dic_{0} = {1}\n'.format(app_uid, u_dic['arr']))
            list_fo.write('tlist_{0} = {1}\n'.format(app_uid, u_dic['arr']))
            c_index += 1
    fo.close()
    fo_edit.close()
    list_fo.close()
    shutil.copy('array_list.py', 'array_listinfo.py')

